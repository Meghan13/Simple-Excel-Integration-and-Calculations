from unittest import TestCase
import unittest

from openpyxl import load_workbook

import ExcelReaderWriter


class TestExcelReaderWriter(TestCase):
    xlrw = ExcelReaderWriter.ExcelReaderWriter("Data_test.xlsx")

    def setUp(self):
        self.xlrw.path = "Data_test.xlsx"
        self.xlrw.work_book = load_workbook(self.xlrw.path)
        self.xlrw.sheet = self.xlrw.work_book.active
        working_copy_path = "test_data_working_copy.xlsx"
        self.xlrw.path = working_copy_path
        self.xlrw.work_book.save(working_copy_path)
        self.xlrw.work_book = load_workbook(working_copy_path)
        self.xlrw.sheet = self.xlrw.work_book.active

    def test_excel_reader(self):
        cell_location = (1,1)
        value = self.xlrw.excel_reader(cell_location)
        assert value == "this is a test"


    def test_excel_writer(self):
        cell_location = (3, 3)
        value = "Testing writer method"
        assert self.xlrw.excel_reader(cell_location) is None
        assert self.xlrw.excel_writer(value, cell_location)
        self.xlrw.excel_writer(value, cell_location)
        test_cell_value = self.xlrw.excel_reader(cell_location)
        self.xlrw.work_book.save("test_data_working_copy.xlsx")
        print(test_cell_value)
        assert test_cell_value == value

    def tearDown(self):
        self.xlrw.work_book.save("test_data_working_copy.xlsx")

