from openpyxl import Workbook, load_workbook


class ExcelReaderWriter:
    # path = "Data.xlsx"
    # work_book = load_workbook(path)
    # sheet = work_book.active

    def __init__(self, file_name="Data.xlsx"):
        self.path = file_name
        self.work_book = load_workbook(self.path)
        self.sheet = self.work_book.active


    def excel_reader(self, cell_location):
        cell = self.sheet.cell(row=cell_location[0], column=cell_location[1])
        return cell.value
        # print(cell.value)

    def excel_writer(self, value, cell_location):
        active_cell = self.sheet.cell(row=cell_location[0], column=cell_location[1])
        active_cell.value = value
        self.work_book.save(self.path)
        return True
